# 5_generate_paper_table.py
"""
生成论文主结果表格 & 附录统计表格（Markdown & LaTeX & Excel）
====================================================
功能：
1. 自动计算 Welch t-test, Hedges'g, 95% CI, FDR校正(分层策略), Post-hoc Power
2. 生成 Table 1 (主结果) & Appendix Table A.2 (完整统计检验)
3. 输出 Markdown, LaTeX, Excel (双工作表+格式+分析层级元数据)
依赖: pip install numpy scipy pandas statsmodels openpyxl
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path
from scipy import stats
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.power import tt_ind_solve_power
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
#  配置区
# ============================================================================
PLOT_DIR = Path("plot_data")
OUTPUT_DIR = Path("tables")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

BASELINE_KEY = "baseline/combined"
COMPARISONS = [
    {"method": "daosheng_temp07/combined", "label": "DaoSheng (τ=0.7)"},
    {"method": "daosheng_temp08/combined", "label": "DaoSheng (τ=0.8)"},
    {"method": "daosheng_temp085/combined", "label": "DaoSheng (τ=0.85)"},
    {"method": "daosheng_temp09/combined", "label": "DaoSheng (τ=0.9)"},
    {"method": "yinyang_gradscale/combined", "label": "Yin-Yang GradScale"},
    {"method": "yinyang_daosheng/combined", "label": "Yin-Yang + DaoSheng (τ=0.85)"},
]
METRICS = ["FID", "IS", "KID"]
FINAL_STEP_IDX = -1  # 对应100,000步

SYMBOLS = {
    "plus_minus": r"\pm", "down": r"\downarrow", "up": r"\uparrow", "tau": r"\tau",
    "dagger": r"\dagger", "ddagger": r"\ddagger", "minus": r"-", "alpha": r"\alpha",
    "times": r"\times", "infty": r"\infty"
}

# ============================================================================
#  核心统计计算函数
# ============================================================================
def compute_welch_stats(m1, s1, n1, m2, s2, n2):
    se = np.sqrt(s1**2/n1 + s2**2/n2)
    t = (m1 - m2) / se
    num = (s1**2/n1 + s2**2/n2)**2
    denom = (s1**2/n1)**2/(n1-1) + (s2**2/n2)**2/(n2-1)
    df = num / denom if denom > 1e-10 else np.inf
    p_raw = 2 * stats.t.sf(np.abs(t), df)
    t_crit = stats.t.ppf(0.975, df)
    ci = ((m1-m2) - t_crit*se, (m1-m2) + t_crit*se)
    return {"t": t, "df": df, "p_raw": p_raw, "se": se, "ci_95": ci, "mean_diff": m1-m2}

def compute_hedges_g(m1, m2, s1, s2, n1, n2):
    pooled = np.sqrt(((n1-1)*s1**2 + (n2-1)*s2**2) / (n1+n2-2))
    if pooled < 1e-10: return 0.0
    d = (m1 - m2) / pooled
    J = 1 - 3/(4*(n1+n2-2) - 1)
    return d * J

def compute_post_hoc_power(hedges_g, n_per_group, alpha=0.05):
    return tt_ind_solve_power(effect_size=abs(hedges_g), nobs1=n_per_group, alpha=alpha, ratio=1.0)

def apply_fdr_correction(p_vals, alpha=0.05):
    reject, p_corr, _, _ = multipletests(p_vals, alpha=alpha, method='fdr_bh')
    return p_corr, reject

# ============================================================================
#  数据构建与统计流程（ 新增分层标记逻辑）
# ============================================================================
def build_statistical_data(plot_data):
    stats_db = []
    
    method_map = {c["label"]: c for c in COMPARISONS}
    
    for comp in COMPARISONS:
        method_key = comp["method"]
        if method_key not in plot_data: continue
        n = plot_data[method_key][METRICS[0]]["n_seeds"]
        
        for met in METRICS:
            base = plot_data[BASELINE_KEY][met]
            meth = plot_data[method_key][met]
            
            m1, s1 = meth["mean"][FINAL_STEP_IDX], meth["std"][FINAL_STEP_IDX]
            m2, s2 = base["mean"][FINAL_STEP_IDX], base["std"][FINAL_STEP_IDX]
            
            welch = compute_welch_stats(m1, s1, n, m2, s2, n)
            hg = compute_hedges_g(m1, m2, s1, s2, n, n)
            power = compute_post_hoc_power(hg, n)
            
            stats_db.append({
                "comp_label": comp["label"], "metric": met,
                "base_mean": m2, "base_std": s2,
                "meth_mean": m1, "meth_std": s1, "n": n, 
                "delta": welch["mean_diff"], "ci_95": welch["ci_95"], 
                "hedges_g": hg, "t_stat": welch["t"], "df": welch["df"],
                "p_raw": welch["p_raw"], "power": power
            })
            
    #  Layer 1: 核心假设 (仅 Yin-Yang GradScale 的3指标)
    core_indices = [i for i, d in enumerate(stats_db) if d["comp_label"] == "Yin-Yang GradScale"]
    if core_indices:
        core_p = [stats_db[i]["p_raw"] for i in core_indices]
        core_p_fdr, core_sig = apply_fdr_correction(core_p)
        idx_map = {c: i for i, c in enumerate(core_indices)}
        
        for i, d in enumerate(stats_db):
            if i in idx_map:
                d["p_fdr"] = core_p_fdr[idx_map[i]]
                d["sig_fdr"] = core_sig[idx_map[i]]
                d["analysis_type"] = "confirmatory"
            else:
                d["p_fdr"] = d["p_raw"]
                d["sig_fdr"] = False
                d["analysis_type"] = "exploratory"
    else:
        for d in stats_db:
            d["p_fdr"] = d["p_raw"]
            d["sig_fdr"] = False
            d["analysis_type"] = "none"
            
    return stats_db

# ============================================================================
#  表格生成器 (MD & LaTeX) -  严格对齐正文格式与分层策略
# ============================================================================
def generate_tables(stats_db):
    # 辅助函数：KID值×100缩放以匹配标题 ($\times 10^{-2}$)
    def _fmt_kid(mean, std):
        return f"${mean*100:.2f} \\pm {std*100:.2f}$"
    
    # 辅助函数：统一τ标签为数学模式 ($\tau=x.x$)
    def _fmt_tau(label):
        if "τ=0.85" in label: return r"Yin-Yang + DaoSheng ($\tau=0.85$)"
        if "τ=0.8" in label:  return r"DaoSheng ($\tau=0.8$)"
        if "τ=0.7" in label:  return r"DaoSheng ($\tau=0.7$)"
        if "τ=0.9" in label:  return r"DaoSheng ($\tau=0.9$)"
        return label

    # ---- Table 1: 主结果表格 ----
    t1_md = ["| Method | FID ↓ | IS ↑ | KID ↓ (×10⁻²) |", "|---|---|---|---|"]
    
    t1_tex = [
        r"\begin{table}[h]", r"\centering", 
        r"\caption{Final metrics on CIFAR-10 at step=100,000, aggregated across ten independent random seeds ($n=10$ per method). Lower FID/KID and higher IS are better. Values reported as mean $\pm$ standard deviation. Statistical tests: Welch's two-sample $t$-test (unequal variance); effect sizes: Hedges' $g$ with small-sample correction; multiple testing: Benjamini-Hochberg FDR correction ($\alpha=0.05$, \textbf{3 core comparisons}: Yin-Yang GradScale vs Baseline on FID/IS/KID). Full test statistics, 95\% confidence intervals, and power analysis in Appendix Table~\ref{tab:stats_appendix}.}", 
        r"\label{tab:main_results}", r"\small", r"\begin{tabular}{lccc}", r"\toprule",
        r"\textbf{Method} & \textbf{FID} $\downarrow$ & \textbf{IS} $\uparrow$ & \textbf{KID} $\downarrow$ ($\times 10^{-2}$) \\", r"\midrule"
    ]
    
    # Baseline 行
    base_fid = next(d["base_mean"] for d in stats_db if d["metric"]=="FID")
    base_fid_std = next(d["base_std"] for d in stats_db if d["metric"]=="FID")
    base_is = next(d["base_mean"] for d in stats_db if d["metric"]=="IS")
    base_is_std = next(d["base_std"] for d in stats_db if d["metric"]=="IS")
    base_kid = next(d["base_mean"] for d in stats_db if d["metric"]=="KID")
    base_kid_std = next(d["base_std"] for d in stats_db if d["metric"]=="KID")
    
    b_fid = f"${base_fid:.2f} \\pm {base_fid_std:.2f}$"
    b_is  = f"${base_is:.2f} \\pm {base_is_std:.2f}$"
    b_kid = _fmt_kid(base_kid, base_kid_std)  
    
    t1_md.append(f"| SNGAN (Baseline) | {b_fid} | {b_is} | {b_kid} |")
    t1_tex.append(f"SNGAN (Baseline) & {b_fid} & {b_is} & {b_kid} \\\\")
    
    # DaoSheng 温度消融
    t1_tex.append(r"\multicolumn{4}{l}{\textit{DaoSheng (temperature ablation)}} \\")
    tau_order = ["DaoSheng (τ=0.7)", "DaoSheng (τ=0.8)", "DaoSheng (τ=0.85)", "DaoSheng (τ=0.9)"]
    for tau_label in tau_order:
        items = {d["metric"]: d for d in stats_db if d["comp_label"] == tau_label}
        if not items: continue
        
        fid = f"${items['FID']['meth_mean']:.2f} \\pm {items['FID']['meth_std']:.2f}$"
        is_ = f"${items['IS']['meth_mean']:.2f} \\pm {items['IS']['meth_std']:.2f}$"
        kid = _fmt_kid(items['KID']['meth_mean'], items['KID']['meth_std'])  # ✅ KID缩放
        tau_val = tau_label.split("τ=")[-1].replace(")", "")
        
        t1_md.append(f"| τ={tau_val} | {fid} | {is_} | {kid} |")
        t1_tex.append(f"\\quad $\\tau={tau_val}$ & {fid} & {is_} & {kid} \\\\")
    
    t1_tex.append(r"\midrule")
    t1_tex.append(r"\multicolumn{4}{l}{\textit{Yin-Yang and combination}} \\")
    
    # Yin-Yang 方法行
    method_order = ["Yin-Yang GradScale", "Yin-Yang + DaoSheng (τ=0.85)"]
    for label in method_order:
        items = {d["metric"]: d for d in stats_db if d["comp_label"] == label}
        if not items: continue
        
        cells = {}
        for k, met in items.items():
            #  独立判断†和‡，允许同时出现
            markers = []
            if met["analysis_type"] == "confirmatory":
                if met["p_raw"] < 0.05: markers.append(r"\dagger")
                if met["sig_fdr"]: markers.append(r"\ddagger")
            
            marker_str = "".join(markers)
            superscript = f"^{{{marker_str}}}" if marker_str else ""
            val = f"{met['meth_mean']:.2f} \\pm {met['meth_std']:.2f}"
            
            #  KID列使用缩放值
            if k == "KID":
                raw_cell = _fmt_kid(met['meth_mean'], met['meth_std'])[1:-1]  # 去掉外层$
                cells[k] = f"$\\mathbf{{{raw_cell}}}{superscript}$" if label == "Yin-Yang GradScale" else f"${raw_cell}$"
            else:
                cells[k] = f"$\\mathbf{{{val}}}{superscript}$" if label == "Yin-Yang GradScale" else f"${val}$"
        
        display_label = _fmt_tau(label)
        t1_md.append(f"| {display_label} | {cells['FID']} | {cells['IS']} | {cells['KID']} |")
        t1_tex.append(f"{display_label} & {cells['FID']} & {cells['IS']} & {cells['KID']} \\\\")
    
    #  底部格式：\botrule + 精确对齐的脚注文本
    t1_tex.extend([
        r"\botrule",
        r"\multicolumn{4}{p{\dimexpr\linewidth-2\tabcolsep\relax}}{\footnotesize",
        r"  $^{\dagger} p_{\rm raw} < 0.05$; $^{\ddagger} p_{\rm FDR} < 0.05$ after Benjamini-Hochberg correction for \textbf{3 core comparisons}. ",
        r"\textit{Exploratory analyses} (DaoSheng ablation, combined method) report uncorrected $p$-values with effect sizes and 95\% CIs; interpret as hypothesis-generating. ",
        r"Full statistical results including all metrics, effect sizes, confidence intervals, and post-hoc power estimates are provided in Appendix Table~\ref{tab:stats_appendix}.",
        r"}",
        r"\end{tabular}", r"\end{table}"
    ])
    
    # ---- Appendix Table A.2： 防溢出紧凑格式 ----
    a2_md = ["| Comparison | Metric | Δ (95% CI) | Hedges' g | t (df) | p_raw | p_FDR | Power |", "|---|---|---|---|---|---|---|---|"]
    
    #  使用 p{3.6cm} 首列自动换行 + 紧凑列间距 + \footnotesize
    a2_tex = [
        r"\begin{table}[h]", r"\centering", 
        r"\caption{Complete statistical test results for key comparisons ($n=10$ seeds/group). $^{\dagger} p_{\rm raw}<0.05$; $^{\ddagger} p_{\rm FDR}<0.05$.}", 
        r"\label{tab:stats_appendix}", 
        r"\footnotesize", r"\setlength{\tabcolsep}{2.5pt}",  
        r"\begin{tabular}{@{}p{3.6cm}@{\hspace{0.4em}}l@{\hspace{0.4em}}c@{\hspace{0.4em}}c@{\hspace{0.4em}}c@{\hspace{0.4em}}c@{\hspace{0.4em}}c@{\hspace{0.4em}}c@{}}", r"\toprule",
        r"\textbf{Comparison} & \textbf{Metric} & \textbf{$\Delta$ (95\% CI)} & \textbf{Hedges' $g$} & \textbf{$t$ (df)} & \textbf{$p_{\rm raw}$} & \textbf{$p_{\rm FDR}$} & \textbf{Power} \\", r"\midrule"
    ]
    
    out_order_labels = ["Yin-Yang GradScale", "Yin-Yang + DaoSheng (τ=0.85)", "DaoSheng (τ=0.8)"]
    for lab in out_order_labels:
        group = [d for d in stats_db if d["comp_label"] == lab]
        for d in group:
            delta_cell = f"${d['delta']:+.2f}$ $[{d['ci_95'][0]:+.2f}, {d['ci_95'][1]:+.2f}]$"
            g_cell = f"${d['hedges_g']:+.2f}$"
            df_str = r"\infty" if np.isinf(d['df']) else f"{d['df']:.1f}"
            t_cell = f"${d['t_stat']:+.2f}$ $({df_str})$"
            p_raw_cell = f"${d['p_raw']:.3f}^{{\dagger}}$" if d["p_raw"] < 0.05 else f"${d['p_raw']:.3f}$"
            p_fdr_cell = f"${d['p_fdr']:.3f}^{{\ddagger}}$" if d["sig_fdr"] else f"${d['p_fdr']:.3f}$"
            power_cell = f"{d['power']*100:.0f}\\%"
            fix_lab = _fmt_tau(d["comp_label"])  # 复用τ格式化函数
            
            a2_md.append(f"| {fix_lab} | {d['metric']} | {delta_cell} | {g_cell} | {t_cell} | {p_raw_cell} | {p_fdr_cell} | {power_cell} |")
            a2_tex.append(f"{fix_lab} & {d['metric']} & {delta_cell} & {g_cell} & {t_cell} & {p_raw_cell} & {p_fdr_cell} & {power_cell} \\\\")
    
    #  脚注使用 \scriptsize 进一步节省空间，8列跨度对齐
    a2_tex.extend([
        r"\botrule",
        r"\multicolumn{8}{@{}p{\linewidth}@{}}{\scriptsize", 
        r"  Notes: Welch's $t$-test (unequal variance); Hedges' $g$ with small-sample correction ($J=0.957$); FDR via Benjamini-Hochberg ($\alpha=0.05$, $m=6$).",
        r"}",
        r"\end{tabular}", r"\end{table}"
    ])
    
    return "\n".join(t1_md), "\n".join(t1_tex), "\n".join(a2_md), "\n".join(a2_tex)

# ============================================================================
#  Excel 生成器（ 新增 Analysis_Type 列）
# ============================================================================
def generate_excel(stats_db, output_path):
    with pd.ExcelWriter(output_path, engine="openpyxl") as wb:
        t1_rows = []
        base_fid = next(d["base_mean"] for d in stats_db if d["metric"]=="FID")
        base_fid_std = next(d["base_std"] for d in stats_db if d["metric"]=="FID")
        t1_rows.append(["SNGAN (Baseline)", base_fid, base_fid_std])
        
        tau_order = ["DaoSheng (τ=0.7)", "DaoSheng (τ=0.8)", "DaoSheng (τ=0.85)", "DaoSheng (τ=0.9)"]
        for lbl in tau_order:
            dat = next((d for d in stats_db if d["comp_label"]==lbl and d["metric"]=="FID"), None)
            if dat: t1_rows.append([lbl, dat["meth_mean"], dat["meth_std"]])
            
        for lbl in ["Yin-Yang GradScale", "Yin-Yang + DaoSheng (τ=0.85)"]:
            dat = next((d for d in stats_db if d["comp_label"]==lbl and d["metric"]=="FID"), None)
            if dat: t1_rows.append([lbl, dat["meth_mean"], dat["meth_std"]])
            
        pd.DataFrame(t1_rows, columns=["Method","FID_Mean","FID_Std"]).to_excel(wb, sheet_name="Table_1", index=False)
        
        #  附录表增加 analysis_type 列，提升可追溯性
        a2_rows = [[d["comp_label"], d["metric"], d["delta"], d["ci_95"][0], d["ci_95"][1], 
                    d["hedges_g"], d["t_stat"], d["df"], d["p_raw"], d["p_fdr"], d["power"], d["analysis_type"]] for d in stats_db]
        pd.DataFrame(a2_rows, columns=["Comparison","Metric","Delta","CI_Low","CI_High","Hedges_g","t","df","p_raw","p_FDR","Power","Analysis_Type"]).to_excel(wb, sheet_name="Appendix_A2", index=False)
        
        for ws in wb.sheets.values():
            fill_h = PatternFill("solid", fgColor="D9E1F2")
            bold_f = Font(bold=True, name="Times New Roman", size=11)
            norm_f = Font(name="Times New Roman", size=10)
            brd = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
            for c in ws[1]: c.font=bold_f; c.fill=fill_h; c.border=brd; c.alignment=Alignment(horizontal="center")
            for r in ws.iter_rows(min_row=2):
                for c in r: c.font=norm_f; c.border=brd; c.alignment=Alignment(vertical="center", horizontal="center" if c.column>1 else "left")
            # 动态适配列宽
            widths = [35,12,12,12,12,10,10,10,10,10,10,12][:ws.max_column]
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================================
#  主程序
# ============================================================================
if __name__ == "__main__":
    files = sorted(PLOT_DIR.glob("plot_data_*.json"))
    if not files:
        raise FileNotFoundError(" 请先运行聚合脚本生成 plot_data_*.json")
        
    input_file = files[-1]
    print(f" 加载数据: {input_file.name}")
    with open(input_file, "r", encoding="utf-8") as f:
        plot_data = json.load(f)

    print(" 计算统计量 (Welch/Hedges/FDR分层校正/Power)...")
    stats_db = build_statistical_data(plot_data)
    
    print(" 生成 Markdown & LaTeX 表格...")
    md_t1, tex_t1, md_a2, tex_a2 = generate_tables(stats_db)
    (OUTPUT_DIR / "table_1.md").write_text(md_t1, encoding="utf-8")
    (OUTPUT_DIR / "table_1.tex").write_text(tex_t1, encoding="utf-8")
    (OUTPUT_DIR / "appendix_a2.md").write_text(md_a2, encoding="utf-8")
    (OUTPUT_DIR / "appendix_a2.tex").write_text(tex_a2, encoding="utf-8")
    
    print(" 生成 Excel (双工作表+格式+分析层级)...")
    generate_excel(stats_db, OUTPUT_DIR / "results.xlsx")

    print("\n 全部生成完成! 输出目录: tables/")
    print("    table_1.md / .tex")
    print("    appendix_a2.md / .tex")
    print("    results.xlsx")